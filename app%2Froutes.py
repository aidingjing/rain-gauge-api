#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
API路由文件
"""

from flask import Blueprint, request, jsonify, send_file, make_response
from datetime import datetime
from app.models import (
    ExceptionData, ST_STBPRP_B, AD_CD_B
)
from app.schemas import (
    PaginationSchema, UpdateRemarkSchema, ResponseSchema,
    ExceptionListResponseSchema, UpdateResponseSchema, FarmListResponseSchema
)
from marshmallow import ValidationError
import logging
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# 创建蓝图
exception_bp = Blueprint('exception', __name__)
logger = logging.getLogger(__name__)


@exception_bp.route('/getExecStationList', methods=['GET'])
def getExecStationList():
    """查询异常数据API - 完全相同的接口"""
    try:
        # 预处理请求参数 - 转换时间格式
        args_dict = request.args.to_dict()

        # 打印调试信息
        logger.info(f"原始参数: {args_dict}")

        # 简单的时间格式转换 (处理多种格式)
        def convert_time_format(time_str):
            if not time_str:
                return time_str

            # 处理 2025-10-10 08:00 格式 (缺少秒数)
            if len(time_str) == 16 and time_str.count('-') == 2 and time_str.count(':') == 1:
                return f"{time_str}:00"

            # 处理 2025-10-10+08:00 格式 (URL编码的空格)
            if '+' in time_str and time_str.count('-') == 2 and len(time_str) == 16:
                date_part = time_str[:10]      # 2025-10-10
                time_part = time_str[11:16]    # 08:00
                return f"{date_part} {time_part}:00"

            # 处理只有日期的格式 2025-10-10
            if len(time_str) == 10 and time_str.count('-') == 2:
                return f"{time_str} 00:00:00"

            return time_str

        # 转换时间参数
        for param in ['bt', 'et', 'start_time', 'end_time']:
            if param in args_dict:
                converted = convert_time_format(args_dict[param])
                if converted != args_dict[param]:
                    logger.info(f"时间参数转换: {param}: {args_dict[param]} -> {converted}")
                args_dict[param] = converted

        logger.info(f"转换后参数: {args_dict}")

        # 使用转换后的参数进行验证
        schema = PaginationSchema()
        try:
            args = schema.load(args_dict)
        except ValidationError as e:
            logger.warning(f"参数验证失败: {e.messages}")
            return jsonify({
                'code': 400,
                'message': '请求参数错误',
                'error': str(e.messages)
            }), 400

        # 解析参数 - 支持新旧参数名，添加URL解码处理
        page = args['page']
        page_size = args['page_size']
        adcd = args.get('adcd') or args.get('aid')
        stcd = args.get('stcd')
        bt = args.get('bt') or args.get('start_time')
        et = args.get('et') or args.get('end_time')

        # 对name参数进行URL解码，处理特殊字符
        name = args.get('name')
        if name:
            from urllib.parse import unquote
            try:
                name = unquote(name, encoding='utf-8')
                logger.info(f"name参数URL解码: {args.get('name')} -> {name}")

                # 检查是否因为#号截断导致不完整
                if len(name) > 0 and not any(char in name for char in ['#', '(', ')', '&', '?']):
                    # 如果name参数看起来可能被截断，记录警告
                    if 'status' not in args:
                        logger.warning(f"status参数缺失，可能因为#号截断问题。当前name: {name}")

            except Exception as e:
                logger.warning(f"name参数URL解码失败: {e}")

        status = args.get('status')
        export = args.get('export')

        # 默认查询待反馈的异常数据（保持向后兼容）
        if status is None:
            status = 0

        # 如果没有指定adcd，则查询所有行政区划（重要修复）
        if adcd is None:
            adcd = ''  # 空字符串会匹配所有记录

        # 如果没有指定时间范围，设置一个很宽的范围（重要修复）
        if bt is None:
            bt = '2020-01-01 00:00:00'
        if et is None:
            et = '2030-12-31 23:59:59'

        logger.info(f"查询异常数据 - 页码: {page}, 每页: {page_size}, 行政区: {adcd}, 时间: {bt} 到 {et}, 状态: {status}")

        # 获取数据库管理器
        from flask import current_app
        db_manager = current_app.config['DATABASE_MANAGER']

        # 调试查询：统计符合基本条件的记录数
        debug_sql = """
        SELECT COUNT(*) as total_count,
               MIN(tm) as earliest_time,
               MAX(tm) as latest_time
        FROM TZX_STCD_EXCE
        WHERE aid LIKE %s
          AND tm >= %s
          AND tm <= %s
          AND rem IS NULL
        """
        debug_params = [f"{adcd}%", bt, et]
        debug_result = db_manager.execute_query(debug_sql, tuple(debug_params))
        if debug_result['success'] and len(debug_result['results']) > 0:
            logger.info(f"调试信息 - 符合条件记录数: {debug_result['results'][0]}")
        else:
            logger.info("调试信息 - 没有找到符合条件的记录")

        # 如果是导出请求，获取所有数据
        logger.info(f"检查导出参数: export={export}")
        if export == 'excel':
            # 将字符串时间转换为datetime对象
            bt_dt = datetime.strptime(bt, '%Y-%m-%d %H:%M:%S') if bt else None
            et_dt = datetime.strptime(et, '%Y-%m-%d %H:%M:%S') if et else None

            all_items = ExceptionData.get_all_filtered_exceptions(
                db_manager, adcd=adcd, stcd=stcd,
                start_time=bt_dt, end_time=et_dt, name=name, status=status
            )

            # 构建导出数据格式
            export_result = {
                'total': len(all_items),
                'page': 1,
                'page_size': len(all_items),
                'pages': 1,
                'items': all_items
            }

            return generateExcelFile(export_result, adcd, stcd, bt_dt, et_dt, name, status)

        # 将字符串时间转换为datetime对象
        if bt is not None:
            if isinstance(bt, str):
                bt_dt = datetime.strptime(bt, '%Y-%m-%d %H:%M:%S')
            elif isinstance(bt, datetime):
                bt_dt = bt
            else:
                logger.warning(f"未知的bt参数类型: {type(bt)}, 值: {bt}")
                bt_dt = None
        else:
            bt_dt = None

        if et is not None:
            if isinstance(et, str):
                et_dt = datetime.strptime(et, '%Y-%m-%d %H:%M:%S')
            elif isinstance(et, datetime):
                et_dt = et
            else:
                logger.warning(f"未知的et参数类型: {type(et)}, 值: {et}")
                et_dt = None
        else:
            et_dt = None

        # 查询分页数据
        logger.info(f"准备调用get_pending_exceptions: page={page}, page_size={page_size}, adcd='{adcd}', status={status}")
        result = ExceptionData.get_pending_exceptions(
            db_manager, page=page, page_size=page_size,
            adcd=adcd, stcd=stcd, start_time=bt_dt, end_time=et_dt, name=name, status=status
        )
        logger.info(f"get_pending_exceptions返回结果: {result}")

        # 构建响应
        response_data = {
            'code': 200,
            'message': '查询成功',
            'data': result
        }

        logger.info(f"查询成功，返回 {len(result['items'])} 条记录")
        return jsonify(response_data)

    except Exception as e:
        logger.error(f"查询异常数据失败: {str(e)}")
        return jsonify({
            'code': 500,
            'message': '服务器内部错误',
            'error': str(e)
        }), 500


@exception_bp.route('/remarkExecInfo', methods=['POST'])
def remarkExecInfo():
    """填写异常原因API - 完全相同的接口"""
    try:
        # 时间格式转换函数 - 与getExecStationList保持一致
        def convert_time_format(time_str):
            if not time_str:
                return time_str

            # 处理 2025-10-10 08:00 格式 (缺少秒数)
            if len(time_str) == 16 and time_str.count('-') == 2 and time_str.count(':') == 1:
                return f"{time_str}:00"

            # 处理 2025-10-10+08:00 格式 (URL编码的空格)
            if '+' in time_str and time_str.count('-') == 2 and len(time_str) == 16:
                date_part = time_str[:10]      # 2025-10-10
                time_part = time_str[11:16]    # 08:00
                return f"{date_part} {time_part}:00"

            # 处理只有日期的格式 2025-10-10
            if len(time_str) == 10 and time_str.count('-') == 2:
                return f"{time_str} 00:00:00"

            return time_str

        # 获取原始数据并进行时间转换
        logger.info(f"请求表单数据: {dict(request.form)}")
        logger.info(f"请求Content-Type: {request.content_type}")
        logger.info(f"请求方法: {request.method}")

        raw_data = {
            'stcd': request.form.get('stcd'),
            'rem': request.form.get('rem'),
            'tm': request.form.get('tm'),
            'name': request.form.get('name'),
            'status': request.form.get('status')
        }

        logger.info(f"解析的原始数据: {raw_data}")

        # 对时间参数进行格式转换
        if raw_data['tm']:
            converted_tm = convert_time_format(raw_data['tm'])
            if converted_tm != raw_data['tm']:
                logger.info(f"时间参数转换: tm: {raw_data['tm']} -> {converted_tm}")
                raw_data['tm'] = converted_tm

        logger.info(f"转换后参数: {raw_data}")

        # 使用转换后的数据进行验证
        schema = UpdateRemarkSchema()
        try:
            validated_data = schema.load(raw_data)
        except ValidationError as e:
            logger.warning(f"参数验证失败: {e.messages}")
            return jsonify({
                'code': 400,
                'message': '请求参数错误',
                'error': str(e.messages)
            }), 400

        # 解析参数
        stcd = validated_data['stcd']
        rem = validated_data['rem']
        tm = validated_data['tm']
        name = validated_data['name']
        status = validated_data['status']

        logger.info(f"更新异常原因 - 测站: {stcd}, 异常时间: {tm}, 反馈人员: {name}, 状态: {status}")

        # 获取数据库管理器
        from flask import current_app
        db_manager = current_app.config['DATABASE_MANAGER']

        # 更新数据库 - 使用pymysql版本
        result = ExceptionData.update_remark(
            db_manager, stcd=stcd, rem=rem, tm=tm,
            re_name=name, status=status
        )

        # 构建响应
        response_data = {
            'code': 200,
            'success': True,
            'message': '异常原因更新成功',
            'data': {
                'station_code': stcd,
                'exception_time': result['detail']['exception_time'],
                'updated_at': result['detail']['updated_at']
            }
        }

        logger.info(f"更新成功，测站: {stcd}, 异常时间: {result['detail']['exception_time']}")
        return jsonify(response_data)

    except ValueError as e:
        logger.warning(f"更新失败: {str(e)}")
        if '没有待反馈的异常数据' in str(e):
            return jsonify({
                'code': 1001,
                'success': False,
                'message': '未找到符合条件的异常数据',
                'error': str(e),
                'suggestion': '请检查测站编码和异常时间是否正确，或该异常数据可能已经被处理过',
                'action': '建议：1. 核对测站编码 2. 确认异常时间精确到秒 3. 检查数据是否已被处理'
            }), 200
        elif '异常时间不匹配' in str(e):
            return jsonify({
                'code': 1002,
                'success': False,
                'message': '异常时间不匹配',
                'error': str(e),
                'suggestion': '请确认要处理的异常时间是否正确，或联系管理员处理',
                'action': '建议：1. 检查时间格式是否为 YYYY-MM-DD HH:MM:SS 2. 确认时间精确到秒'
            }), 200
        elif '未找到匹配的异常记录' in str(e):
            return jsonify({
                'code': 1003,
                'success': False,
                'message': '未找到匹配的异常记录',
                'error': str(e),
                'suggestion': '请检查异常时间是否准确，或查看可用异常时间列表',
                'action': '建议：1. 确认测站编码正确 2. 检查异常时间是否准确 3. 确认数据存在且未处理'
            }), 200
        else:
            return jsonify({
                'code': 400,
                'success': False,
                'message': '业务逻辑错误',
                'error': str(e)
            }), 200

    except Exception as e:
        logger.error(f"更新异常原因失败: {str(e)}")
        return jsonify({
            'code': 500,
            'message': '服务器内部错误',
            'error': str(e)
        }), 500


@exception_bp.route('/farms', methods=['GET'])
def getFarmList():
    """获取团场列表API - 完全相同的接口"""
    try:
        logger.info("获取团场列表")

        # 获取数据库管理器
        from flask import current_app
        db_manager = current_app.config['DATABASE_MANAGER']

        # 查询团场数据 - 使用AD_CD_B类
        farms = AD_CD_B.get_all_regions(db_manager)

        # 构建响应
        response_data = {
            'code': 200,
            'message': '查询成功',
            'data': farms
        }

        logger.info(f"查询成功，返回 {len(farms)} 个团场")
        return jsonify(response_data)

    except Exception as e:
        logger.error(f"获取团场列表失败: {str(e)}")
        return jsonify({
            'code': 500,
            'message': '服务器内部错误',
            'error': str(e)
        }), 500


@exception_bp.route('/exception-data/statistics', methods=['GET'])
def getStatistics():
    """获取异常数据统计信息API - 完全相同的接口"""
    try:
        logger.info("获取异常数据统计信息")

        # 获取数据库管理器
        from flask import current_app
        db_manager = current_app.config['DATABASE_MANAGER']

        # 获取统计信息
        stats = ExceptionData.get_statistics(db_manager)

        # 构建响应
        response_data = {
            'code': 200,
            'message': '查询成功',
            'data': stats
        }

        logger.info(f"统计完成 - 待反馈总数: {stats['total_pending']}, 测站数: {stats['station_count']}, 团场数: {stats['farm_count']}")
        return jsonify(response_data)

    except Exception as e:
        logger.error(f"获取统计信息失败: {str(e)}")
        return jsonify({
            'code': 500,
            'message': '服务器内部错误',
            'error': str(e)
        }), 500


@exception_bp.route('/health', methods=['GET'])
def healthCheck():
    """健康检查API - 完全相同的接口"""
    try:
        logger.info("执行健康检查")

        # 获取数据库管理器
        from flask import current_app
        db_manager = current_app.config['DATABASE_MANAGER']

        # 执行健康检查
        health_result = ExceptionData.health_check(db_manager)

        # 构建响应
        response_data = {
            'code': 200,
            'message': 'API服务正常运行',
            'data': health_result
        }

        logger.info(f"健康检查完成 - 状态: {health_result['status']}")
        return jsonify(response_data)

    except Exception as e:
        logger.error(f"健康检查失败: {str(e)}")
        return jsonify({
            'code': 500,
            'message': '健康检查失败',
            'error': str(e)
        }), 500


def generateExcelFile(result, adcd=None, stcd=None, bt=None, et=None, name=None, status=None):
    """生成Excel文件用于异常数据导出 - 完全相同的函数"""
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "异常数据"

        # 定义列标题
        headers = [
            '测站编码', '测站名称', '行政区代码', '异常值', '异常时间',
            '插入时间', '异常原因', '反馈人员', '处理状态', '反馈时间',
            '经度', '纬度', '县', '市'
        ]

        # 添加标题行
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 填充数据
        for row_num, item in enumerate(result['items'], 2):
            # 状态文本映射
            status_text = '待反馈' if item.get('status') == 0 else '已处理' if item.get('status') == 1 else '未知状态'

            data_row = [
                item.get('stcd', ''),
                item.get('stnm', ''),
                item.get('aid', ''),
                item.get('val', 0),
                item.get('tm', ''),
                item.get('insert_tm', ''),
                item.get('rem', ''),
                item.get('re_name', ''),
                status_text,
                item.get('re_time', ''),
                item.get('lgtd', ''),
                item.get('lttd', ''),
                item.get('xian', ''),
                item.get('shi', '')
            ]

            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 添加筛选条件信息
        ws_info = wb.create_sheet(title="导出信息")
        info_data = [
            ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['总记录数', result.get('total', 0)],
            ['当前页记录数', len(result.get('items', []))],
            ['政区代码', adcd or '全部'],
            ['测站编码', stcd or '全部'],
            ['测站名称', name or '全部'],
            ['起始时间', bt.strftime('%Y-%m-%d %H:%M:%S') if bt else '全部'],
            ['终止时间', et.strftime('%Y-%m-%d %H:%M:%S') if et else '全部'],
            ['状态', _getStatusText(status)]
        ]

        for row_num, (label, value) in enumerate(info_data, 1):
            ws_info.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            ws_info.cell(row=row_num, column=2, value=value)

        # 将工作簿保存到内存中的字节流
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # 生成文件名
        filename = f"异常数据导出_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # 记录导出日志
        logger.info(f"Excel文件导出成功 - 文件名: {filename}, 记录数: {len(result.get('items', []))}")

        # 返回文件下载响应
        response = make_response(
            send_file(
                excel_file,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        )

        # 添加响应头
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

        return response

    except Exception as e:
        logger.error(f"生成Excel文件失败: {str(e)}")
        return jsonify({
            'code': 500,
            'message': 'Excel文件生成失败',
            'error': str(e)
        }), 500


def _getStatusText(status):
    """获取状态文本 - 完全相同的函数"""
    if status is None:
        return '待反馈'
    elif status == 0:
        return '待反馈'
    elif status == 1:
        return '已处理'
    elif status == 2:
        return '所有记录'
    else:
        return '未知状态'


